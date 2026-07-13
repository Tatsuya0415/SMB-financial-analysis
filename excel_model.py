"""
Excel財務モデル生成モジュール
6シート構成: 財務3表 / 経営指標 / 趨勢分析 / 業界比較 / シナリオ / 資金繰り
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import copy


# ─────────────────────────────────────────
# カラーパレット（コンサル風ダークネイビー＋アクセント）
# ─────────────────────────────────────────
C = {
    "navy":      "1B3A5C",
    "navy_mid":  "2C5282",
    "blue":      "3182CE",
    "light_blue":"EBF8FF",
    "green":     "276749",
    "green_bg":  "C6F6D5",
    "amber":     "B7791F",
    "amber_bg":  "FEFCBF",
    "red":       "C53030",
    "red_bg":    "FED7D7",
    "white":     "FFFFFF",
    "gray_bg":   "F7FAFC",
    "gray_mid":  "E2E8F0",
    "gray_dark": "4A5568",
    "black":     "1A202C",
}

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="1A202C", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="游ゴシック")

def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style="thin") -> Border:
    s = Side(style=style, color="CBD5E0")
    return Border(left=s, right=s, top=s, bottom=s)

def _header(ws, row, col, text, width_hint=None):
    """ヘッダーセルスタイル"""
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(C["navy"])
    c.font = _font(bold=True, color=C["white"], size=10)
    c.alignment = _align()
    c.border = _border()
    return c

def _subheader(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(C["navy_mid"])
    c.font = _font(bold=True, color=C["white"], size=9)
    c.alignment = _align()
    c.border = _border()
    return c

def _section_title(ws, row, col, text, span=10):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(C["blue"])
    c.font = _font(bold=True, color=C["white"], size=11)
    c.alignment = _align(h="left")
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    return c

def _kpi_cell(ws, row, col, value, rag_color="white", number_format="#,##0"):
    c = ws.cell(row=row, column=col, value=value)
    bg = {"GREEN": C["green_bg"], "AMBER": C["amber_bg"],
          "RED": C["red_bg"]}.get(rag_color, C["white"])
    c.fill = _fill(bg)
    fc = {"GREEN": C["green"], "AMBER": C["amber"],
          "RED": C["red"]}.get(rag_color, C["black"])
    c.font = _font(bold=True, color=fc, size=10)
    c.alignment = _align()
    c.number_format = number_format
    c.border = _border()
    return c

def _data_row(ws, row, col, label, value, sub=False, number_format="#,##0"):
    lc = ws.cell(row=row, column=col, value=label)
    lc.fill = _fill(C["gray_bg"] if not sub else C["white"])
    lc.font = _font(bold=not sub, size=9, color=C["gray_dark"] if sub else C["black"])
    lc.alignment = _align(h="left")
    lc.border = _border()
    if sub:
        lc.alignment = Alignment(horizontal="left", vertical="center",
                                  indent=2, wrap_text=False)

    vc = ws.cell(row=row, column=col + 1, value=value)
    vc.fill = _fill(C["white"])
    vc.font = _font(size=10)
    vc.alignment = _align(h="right")
    vc.number_format = number_format
    vc.border = _border()
    return lc, vc

# ─────────────────────────────────────────
# Sheet 1 — 財務3表
# ─────────────────────────────────────────
def build_financials_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.active
    ws.title = "①財務3表"
    ws.sheet_view.showGridLines = False

    # 列幅
    col_widths = [2, 28, 16, 16, 16, 2, 28, 16, 16, 16, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    pl = data.get("pl", {})
    bs = data.get("bs", {})
    cf = data.get("cf", {})
    fy = data.get("fiscal_year", "当期")
    fy_prev = data.get("fiscal_year_prev", "前期")

    # ── PL ──
    r = 2
    _section_title(ws, r, 2, "  損益計算書（P/L）", 4)
    r += 1
    for col, txt in zip([2, 3, 4, 5], ["科目", fy, fy_prev, "増減"]):
        _header(ws, r, col, txt)
    r += 1

    pl_items = [
        ("売上高",          pl.get("revenue"),          pl.get("revenue_prev"),          True),
        ("  売上原価",       pl.get("cogs"),             pl.get("cogs_prev"),             False),
        ("売上総利益",       pl.get("gross_profit"),     pl.get("gross_profit_prev"),     True),
        ("  販売費及び一般管理費", pl.get("sga"),         pl.get("sga_prev"),              False),
        ("営業利益",         pl.get("operating_profit"), pl.get("operating_profit_prev"), True),
        ("  営業外収益",     pl.get("non_op_income"),    None,                            False),
        ("  営業外費用",     pl.get("non_op_expense"),   None,                            False),
        ("経常利益",         pl.get("ordinary_profit"),  pl.get("ordinary_profit_prev"),  True),
        ("  特別損益（純額）",pl.get("extraordinary"),   None,                            False),
        ("税引前当期純利益", pl.get("pretax_profit"),    None,                            True),
        ("  法人税等",       pl.get("tax"),              None,                            False),
        ("当期純利益",       pl.get("net_profit"),       pl.get("net_profit_prev"),       True),
    ]

    for label, cur, prev, bold in pl_items:
        sub = not bold
        lc = ws.cell(row=r, column=2, value=label)
        lc.fill = _fill(C["gray_bg"] if bold else C["white"])
        lc.font = _font(bold=bold, size=9,
                         color=C["black"] if bold else C["gray_dark"])
        lc.alignment = _align(h="left")
        lc.border = _border()
        if sub:
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)

        for col, val in zip([3, 4], [cur, prev]):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = _fill(C["white"])
            c.font = _font(bold=bold, size=10)
            c.alignment = _align(h="right")
            c.number_format = "#,##0"
            c.border = _border()

        # 増減
        diff_c = ws.cell(row=r, column=5)
        if cur is not None and prev is not None:
            diff = cur - prev
            diff_c.value = diff
            diff_c.font = _font(bold=bold, size=10,
                                  color=C["green"] if diff >= 0 else C["red"])
            diff_c.number_format = '+#,##0;-#,##0;0'
        diff_c.fill = _fill(C["white"])
        diff_c.alignment = _align(h="right")
        diff_c.border = _border()
        r += 1

    # ── BS ──
    r_bs_start = 2
    _section_title(ws, r_bs_start, 7, "  貸借対照表（B/S）", 4)
    r_bs = r_bs_start + 1
    for col, txt in zip([7, 8, 9, 10], ["科目", fy, fy_prev, "増減"]):
        _header(ws, r_bs, col, txt)
    r_bs += 1

    bs_items = [
        ("【資産の部】",    None, None, True),
        ("  流動資産 計",   bs.get("current_assets"),   bs.get("current_assets_prev"),  True),
        ("    現預金",       bs.get("cash"),             None, False),
        ("    売掛金",       bs.get("receivables"),      None, False),
        ("    棚卸資産",     bs.get("inventory"),        None, False),
        ("    その他流動",   bs.get("other_current"),    None, False),
        ("  固定資産 計",   bs.get("fixed_assets"),     bs.get("fixed_assets_prev"),    True),
        ("    有形固定資産", bs.get("tangible_assets"),  None, False),
        ("    無形固定資産", bs.get("intangible_assets"),None, False),
        ("    投資その他",   bs.get("investments"),      None, False),
        ("総資産",           bs.get("total_assets"),     bs.get("total_assets_prev"),    True),
        ("【負債の部】",    None, None, True),
        ("  流動負債 計",   bs.get("current_liabilities"),bs.get("current_liabilities_prev"), True),
        ("    買掛金",       bs.get("payables"),         None, False),
        ("    短期借入金",   bs.get("short_term_debt"),  None, False),
        ("    その他流動負債",bs.get("other_current_liab"),None, False),
        ("  固定負債 計",   bs.get("fixed_liabilities"), bs.get("fixed_liabilities_prev"), True),
        ("    長期借入金",   bs.get("long_term_debt"),   None, False),
        ("負債合計",         bs.get("total_liabilities"),None, True),
        ("【純資産の部】",  None, None, True),
        ("  資本金",         bs.get("capital"),          None, False),
        ("  利益剰余金",     bs.get("retained_earnings"),None, False),
        ("純資産合計",       bs.get("equity"),           bs.get("equity_prev"),          True),
        ("負債・純資産合計", bs.get("total_assets"),     None, True),
    ]

    for label, cur, prev, bold in bs_items:
        lc = ws.cell(row=r_bs, column=7, value=label)
        is_section = label.startswith("【")
        lc.fill = _fill(C["navy_mid"] if is_section else
                         C["gray_bg"] if bold else C["white"])
        lc.font = _font(bold=bold or is_section, size=9,
                         color=C["white"] if is_section else
                         C["black"] if bold else C["gray_dark"])
        lc.alignment = _align(h="left")
        lc.border = _border()
        if not bold and not is_section:
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=3)

        for col, val in zip([8, 9], [cur, prev]):
            c = ws.cell(row=r_bs, column=col, value=val)
            c.fill = _fill(C["white"])
            c.font = _font(bold=bold, size=10)
            c.alignment = _align(h="right")
            c.number_format = "#,##0"
            c.border = _border()

        diff_c = ws.cell(row=r_bs, column=10)
        if cur is not None and prev is not None:
            diff = cur - prev
            diff_c.value = diff
            diff_c.font = _font(bold=bold, size=10,
                                  color=C["green"] if diff >= 0 else C["red"])
            diff_c.number_format = '+#,##0;-#,##0;0'
        diff_c.fill = _fill(C["white"])
        diff_c.alignment = _align(h="right")
        diff_c.border = _border()
        r_bs += 1

    # ── CF ──
    r_cf = max(r, r_bs) + 1
    _section_title(ws, r_cf, 2, "  キャッシュフロー計算書（C/F）", 9)
    r_cf += 1
    for col, txt in zip([2, 3, 4], ["区分", fy, fy_prev]):
        _header(ws, r_cf, col, txt)
    r_cf += 1

    cf_items = [
        ("営業活動によるCF",  cf.get("operating_cf"),  cf.get("operating_cf_prev"),  True),
        ("投資活動によるCF",  cf.get("investing_cf"),  cf.get("investing_cf_prev"),  True),
        ("財務活動によるCF",  cf.get("financing_cf"),  cf.get("financing_cf_prev"),  True),
        ("現金及び現金同等物の増減", cf.get("net_change"), None, True),
        ("期末現金及び現金同等物", cf.get("ending_cash"), None, True),
    ]

    for label, cur, prev, bold in cf_items:
        lc = ws.cell(row=r_cf, column=2, value=label)
        lc.fill = _fill(C["gray_bg"])
        lc.font = _font(bold=bold, size=9)
        lc.alignment = _align(h="left")
        lc.border = _border()

        for col, val in zip([3, 4], [cur, prev]):
            c = ws.cell(row=r_cf, column=col, value=val)
            c.fill = _fill(C["white"])
            c.font = _font(size=10,
                            color=C["green"] if (val or 0) > 0 else
                            C["red"] if (val or 0) < 0 else C["black"])
            c.alignment = _align(h="right")
            c.number_format = "+#,##0;-#,##0;0"
            c.border = _border()
        r_cf += 1

    ws.freeze_panes = "C4"


# ─────────────────────────────────────────
# Sheet 2 — 経営指標ダッシュボード
# ─────────────────────────────────────────
def build_kpi_sheet(wb: openpyxl.Workbook, ratios: dict, benchmark: dict):
    ws = wb.create_sheet("②経営指標")
    ws.sheet_view.showGridLines = False

    col_widths = [2, 24, 14, 14, 14, 10, 20, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 2
    ws.row_dimensions[r].height = 30

    # タイトル行
    tc = ws.cell(row=r, column=2, value="経営指標ダッシュボード（中小企業診断士基準）")
    tc.font = _font(bold=True, size=14, color=C["navy"])
    tc.alignment = _align(h="left")
    ws.merge_cells(f"B{r}:G{r}")
    r += 1

    # ヘッダー
    for col, txt in zip(range(2, 8), ["指標名", "当社値", "業界平均", "過不足", "評価", "コメント"]):
        _header(ws, r, col, txt)
    r += 1

    from industry_benchmarks import rag_signal, RAG_THRESHOLDS

    kpi_groups = [
        ("収益性指標", [
            ("gross_margin",      "売上総利益率",    "{:.1%}",   False),
            ("operating_margin",  "営業利益率",      "{:.1%}",   False),
            ("ordinary_margin",   "経常利益率",      "{:.1%}",   False),
            ("roa",               "ROA（総資産利益率）", "{:.1%}", False),
            ("roe",               "ROE（自己資本利益率）", "{:.1%}", False),
            ("ebitda_margin",     "EBITDAマージン",  "{:.1%}",   False),
        ]),
        ("安全性指標", [
            ("current_ratio",     "流動比率",        "{:.2f}倍", False),
            ("quick_ratio",       "当座比率",        "{:.2f}倍", False),
            ("equity_ratio",      "自己資本比率",    "{:.1%}",   False),
            ("debt_ratio",        "負債比率（有利子）","{:.2f}倍", True),
            ("interest_coverage", "インタレストカバレッジ","{:.1f}倍", False),
            ("net_debt_ebitda",   "純有利子負債/EBITDA","{:.2f}倍", True),
        ]),
        ("効率性指標", [
            ("asset_turnover",    "総資産回転率",    "{:.2f}回", False),
            ("receivable_days",   "売掛金回転日数",  "{:.0f}日", True),
            ("inventory_days",    "棚卸資産回転日数","{:.0f}日", True),
            ("payable_days",      "買掛金回転日数",  "{:.0f}日", False),
            ("ccc",               "CCC（現金転換サイクル）","{:.0f}日", True),
        ]),
        ("成長性指標", [
            ("revenue_growth",    "売上高成長率",    "{:+.1%}",  False),
            ("op_profit_growth",  "営業利益成長率",  "{:+.1%}",  False),
            ("equity_growth",     "自己資本成長率",  "{:+.1%}",  False),
        ]),
    ]

    for group_name, items in kpi_groups:
        _section_title(ws, r, 2, f"  ▌ {group_name}", 6)
        r += 1
        ws.row_dimensions[r - 1].height = 22

        for key, label, fmt, inverse in items:
            val = ratios.get(key)
            bm_val = benchmark.get(key)
            thresh = RAG_THRESHOLDS.get(key, {})

            # 指標名
            lc = ws.cell(row=r, column=2, value=label)
            lc.fill = _fill(C["white"])
            lc.font = _font(size=9)
            lc.alignment = _align(h="left")
            lc.border = _border()

            # 当社値
            if val is not None:
                try:
                    display = fmt.format(val)
                except:
                    display = str(val)
                vc = ws.cell(row=r, column=3, value=display)
                vc.fill = _fill(C["white"])
                vc.font = _font(size=10, bold=True)
                vc.alignment = _align()
                vc.border = _border()
            else:
                ws.cell(row=r, column=3, value="─").border = _border()

            # 業界平均
            if bm_val is not None:
                try:
                    bm_display = fmt.format(bm_val)
                except:
                    bm_display = str(bm_val)
                bc = ws.cell(row=r, column=4, value=bm_display)
            else:
                bc = ws.cell(row=r, column=4, value="─")
            bc.fill = _fill(C["gray_bg"])
            bc.font = _font(size=9, color=C["gray_dark"])
            bc.alignment = _align()
            bc.border = _border()

            # 過不足
            diff_c = ws.cell(row=r, column=5)
            if val is not None and bm_val is not None:
                diff = val - bm_val
                try:
                    diff_display = fmt.format(abs(diff))
                    sign = "+" if diff > 0 else "-"
                    diff_c.value = f"{sign}{diff_display}"
                except:
                    diff_c.value = f"{diff:+.3f}"
                ok = (diff > 0) != inverse
                diff_c.font = _font(size=9, bold=True,
                                      color=C["green"] if ok else C["red"])
            diff_c.fill = _fill(C["white"])
            diff_c.alignment = _align()
            diff_c.border = _border()

            # RAG評価
            rag = "─"
            rag_color = "white"
            if val is not None and thresh:
                rag = rag_signal(val, thresh["low"], thresh["high"], thresh["inverse"])
                rag_color = rag
                rag_labels = {"GREEN": "✓ 良好", "AMBER": "△ 要注意", "RED": "✕ 要改善"}
                rag = rag_labels.get(rag, rag)
            ec = ws.cell(row=r, column=6, value=rag)
            bg_map = {"white": C["white"], "GREEN": C["green_bg"],
                      "AMBER": C["amber_bg"], "RED": C["red_bg"]}
            fc_map = {"white": C["black"], "GREEN": C["green"],
                      "AMBER": C["amber"], "RED": C["red"]}
            ec.fill = _fill(bg_map.get(rag_color, C["white"]))
            ec.font = _font(bold=True, size=9, color=fc_map.get(rag_color, C["black"]))
            ec.alignment = _align()
            ec.border = _border()

            # コメント自動生成
            comment = _auto_comment(key, val, bm_val, inverse)
            cc = ws.cell(row=r, column=7, value=comment)
            cc.fill = _fill(C["gray_bg"])
            cc.font = _font(size=8, color=C["gray_dark"], italic=True)
            cc.alignment = _align(h="left", wrap=True)
            cc.border = _border()

            r += 1

    ws.freeze_panes = "C4"


def _auto_comment(key: str, val, bm, inverse: bool) -> str:
    if val is None:
        return "データなし"
    if bm is None:
        return ""
    diff_pct = abs(val - bm) / max(abs(bm), 0.001)
    better = (val > bm) != inverse
    degree = "大幅に" if diff_pct > 0.3 else "やや"

    comments = {
        "gross_margin":      ("粗利率が業界平均を上回り価格競争力あり", "粗利率改善が優先課題。値上げ交渉or原価低減を検討"),
        "operating_margin":  ("営業効率良好。販管費管理が奏功", "販管費の見直し・固定費削減が急務"),
        "roa":               ("資産効率が高く経営の質が良い", "遊休資産の整理・資産回転改善が必要"),
        "roe":               ("株主への還元効率が高い", "収益力強化または財務レバレッジ活用を検討"),
        "current_ratio":     ("短期的な支払能力に余裕あり", "短期資金繰りに要注意。流動資産の充実を"),
        "equity_ratio":      ("財務基盤が安定。外部借入余力あり", "自己資本の拡充が財務健全化の鍵"),
        "debt_ratio":        ("有利子負債が低水準。借入余力大", "有利子負債が重い。返済計画の再検討を"),
        "asset_turnover":    ("資産を効率的に活用し売上創出", "資産の稼働率改善・売上拡大が必要"),
        "receivable_days":   ("回収が早く資金効率が良い", "回収サイト長期化リスク。与信管理強化を"),
        "inventory_days":    ("在庫効率が良く資金固定が少ない", "過剰在庫気味。適正在庫管理の導入を"),
        "ccc":               ("現金転換サイクルが短く運転資本効率が高い", "CCCが長く運転資金需要が大きい"),
        "revenue_growth":    ("成長軌道にある", "売上が横ばい〜減少傾向。新規顧客開拓を"),
        "ebitda_margin":     ("キャッシュ創出力が高い", "EBITDA改善のため収益・コスト両面のテコ入れを"),
        "interest_coverage": ("利息支払能力に余裕あり", "利払い負担が重い。借入条件の見直しを"),
        "net_debt_ebitda":   ("債務返済能力が高い", "純有利子負債/EBITDAが高く財務リスクに要注意"),
    }
    pos, neg = comments.get(key, ("良好", "要改善"))
    return f"{degree}{pos if better else neg}"


# ─────────────────────────────────────────
# Sheet 3 — 趨勢分析（折れ線グラフ）
# ─────────────────────────────────────────
def build_trend_sheet(wb: openpyxl.Workbook, trend_data: list):
    """
    trend_data: [{"year": "2020", "revenue": ..., "op_profit": ..., ...}, ...]
    """
    ws = wb.create_sheet("③趨勢分析")
    ws.sheet_view.showGridLines = False

    col_widths = [2, 16, 14, 14, 14, 14, 14, 14, 14, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if not trend_data:
        ws.cell(row=2, column=2, value="複数期データがありません")
        return

    headers = ["期", "売上高", "売上総利益", "営業利益", "経常利益", "純利益",
               "営業利益率", "自己資本比率", "ROA"]
    for col, h in enumerate(headers, 2):
        _header(ws, 2, col, h)

    for i, d in enumerate(trend_data):
        r = 3 + i
        row_vals = [
            d.get("year"),
            d.get("revenue"),
            d.get("gross_profit"),
            d.get("operating_profit"),
            d.get("ordinary_profit"),
            d.get("net_profit"),
            d.get("operating_margin"),
            d.get("equity_ratio"),
            d.get("roa"),
        ]
        for col, v in enumerate(row_vals, 2):
            c = ws.cell(row=r, column=col, value=v)
            c.border = _border()
            c.alignment = _align()
            if col in [3, 4, 5, 6, 7]:
                c.number_format = "#,##0"
            elif col in [8, 9, 10]:
                c.number_format = "0.0%"

    # 折れ線チャート（売上・利益）
    n = len(trend_data)
    chart1 = LineChart()
    chart1.title = "売上高・利益推移"
    chart1.style = 10
    chart1.y_axis.title = "金額（千円）"
    chart1.x_axis.title = "期"
    chart1.width = 20
    chart1.height = 12

    for col_idx, label in [(3, "売上高"), (5, "営業利益"), (6, "経常利益")]:
        data_ref = Reference(ws, min_col=col_idx, min_row=2,
                              max_row=2 + n)
        chart1.add_data(data_ref, titles_from_data=True)

    cats = Reference(ws, min_col=2, min_row=3, max_row=2 + n)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "B" + str(5 + n))

    # 折れ線チャート（利益率）
    chart2 = LineChart()
    chart2.title = "利益率・財務指標推移"
    chart2.style = 10
    chart2.y_axis.numFmt = "0%"
    chart2.y_axis.title = "比率"
    chart2.width = 20
    chart2.height = 12

    for col_idx, label in [(8, "営業利益率"), (9, "自己資本比率"), (10, "ROA")]:
        data_ref = Reference(ws, min_col=col_idx, min_row=2,
                              max_row=2 + n)
        chart2.add_data(data_ref, titles_from_data=True)

    chart2.set_categories(cats)
    ws.add_chart(chart2, "L" + str(5 + n))


# ─────────────────────────────────────────
# Sheet 4 — 業界比較（棒グラフ）
# ─────────────────────────────────────────
def build_benchmark_sheet(wb: openpyxl.Workbook, ratios: dict, benchmark: dict):
    ws = wb.create_sheet("④業界比較")
    ws.sheet_view.showGridLines = False

    col_widths = [2, 22, 14, 14, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    compare_items = [
        ("営業利益率",    "operating_margin",  "{:.1%}"),
        ("経常利益率",    "ordinary_margin",   "{:.1%}"),
        ("ROA",          "roa",               "{:.1%}"),
        ("ROE",          "roe",               "{:.1%}"),
        ("自己資本比率", "equity_ratio",       "{:.1%}"),
        ("流動比率",      "current_ratio",     "{:.2f}倍"),
        ("総資産回転率", "asset_turnover",     "{:.2f}回"),
        ("EBITDAマージン","ebitda_margin",     "{:.1%}"),
    ]

    for col, txt in zip(range(2, 5), ["指標", "当社", benchmark.get("label", "業界平均")]):
        _header(ws, 2, col, txt)

    for i, (label, key, fmt) in enumerate(compare_items):
        r = 3 + i
        lc = ws.cell(row=r, column=2, value=label)
        lc.border = _border()
        lc.font = _font(size=9)
        lc.alignment = _align(h="left")
        lc.fill = _fill(C["gray_bg"] if i % 2 == 0 else C["white"])

        for col, src in zip([3, 4], [ratios, benchmark]):
            val = src.get(key)
            c = ws.cell(row=r, column=col)
            if val is not None:
                try:
                    c.value = fmt.format(val)
                except:
                    c.value = val
            else:
                c.value = "─"
            c.border = _border()
            c.alignment = _align()
            c.fill = _fill(C["gray_bg"] if i % 2 == 0 else C["white"])

    # 棒グラフ
    chart = BarChart()
    chart.type = "col"
    chart.title = f"業界比較（当社 vs {benchmark.get('label', '業界平均')}）"
    chart.style = 10
    chart.grouping = "clustered"
    chart.width = 24
    chart.height = 14

    data_ref = Reference(ws, min_col=3, max_col=4,
                          min_row=2, max_row=2 + len(compare_items))
    chart.add_data(data_ref, titles_from_data=True)
    cats = Reference(ws, min_col=2, min_row=3,
                     max_row=2 + len(compare_items))
    chart.set_categories(cats)
    ws.add_chart(chart, "B" + str(5 + len(compare_items)))


# ─────────────────────────────────────────
# Sheet 5 — シナリオ分析
# ─────────────────────────────────────────
def build_scenario_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("⑤シナリオ分析")
    ws.sheet_view.showGridLines = False

    col_widths = [2, 24, 16, 16, 16, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    pl = data.get("pl", {})
    rev = pl.get("revenue", 0) or 0
    gp = pl.get("gross_profit", 0) or 0
    sga = pl.get("sga", 0) or 0
    op = pl.get("operating_profit", 0) or 0

    _section_title(ws, 2, 2, "  シナリオ分析（P/L 3ケース）", 4)

    r = 3
    for col, txt in zip(range(2, 6), ["科目", "悲観シナリオ", "基本シナリオ", "楽観シナリオ"]):
        hc = _header(ws, r, col, txt)
    r += 1

    # 前提
    scenarios = {
        "悲観": {"rev_chg": -0.10, "gm_chg": -0.02, "sga_chg": 0.05,  "fill": C["red_bg"],   "fc": C["red"]},
        "基本": {"rev_chg":  0.03, "gm_chg":  0.00, "sga_chg": 0.02,  "fill": C["amber_bg"], "fc": C["amber"]},
        "楽観": {"rev_chg":  0.08, "gm_chg":  0.02, "sga_chg": 0.00,  "fill": C["green_bg"], "fc": C["green"]},
    }

    # 前提行
    assump_items = [
        ("売上高成長率前提",  "rev_chg",  "{:+.0%}"),
        ("粗利率変化前提",    "gm_chg",   "{:+.1%}"),
        ("販管費増加率前提",  "sga_chg",  "{:+.0%}"),
    ]
    _subheader(ws, r, 2, "▌ 前提条件")
    ws.merge_cells(f"B{r}:E{r}")
    r += 1

    for key, k, fmt in assump_items:
        lc = ws.cell(row=r, column=2, value=key)
        lc.fill = _fill(C["gray_bg"])
        lc.font = _font(size=9)
        lc.alignment = _align(h="left")
        lc.border = _border()

        for col, (sc_name, sc) in enumerate(scenarios.items(), 3):
            c = ws.cell(row=r, column=col, value=fmt.format(sc[k]))
            c.fill = _fill(sc["fill"])
            c.font = _font(size=10, color=sc["fc"], bold=True)
            c.alignment = _align()
            c.border = _border()
        r += 1

    # P/L シミュレーション
    r += 1
    _subheader(ws, r, 2, "▌ P/Lシミュレーション（千円）")
    ws.merge_cells(f"B{r}:E{r}")
    r += 1

    gm_rate = gp / rev if rev else 0

    for sc_name, sc in scenarios.items():
        sc["rev_sim"]   = int(rev * (1 + sc["rev_chg"]))
        sc["gp_sim"]    = int(sc["rev_sim"] * (gm_rate + sc["gm_chg"]))
        sc["sga_sim"]   = int(sga * (1 + sc["sga_chg"]))
        sc["op_sim"]    = sc["gp_sim"] - sc["sga_sim"]
        sc["op_margin_sim"] = sc["op_sim"] / sc["rev_sim"] if sc["rev_sim"] else 0

    sim_items = [
        ("売上高",      [sc["rev_sim"] for sc in scenarios.values()]),
        ("売上総利益",  [sc["gp_sim"]  for sc in scenarios.values()]),
        ("販管費",      [sc["sga_sim"] for sc in scenarios.values()]),
        ("営業利益",    [sc["op_sim"]  for sc in scenarios.values()]),
        ("営業利益率",  [sc["op_margin_sim"] for sc in scenarios.values()]),
    ]

    for label, vals in sim_items:
        is_margin = "率" in label
        lc = ws.cell(row=r, column=2, value=label)
        lc.fill = _fill(C["gray_bg"])
        lc.font = _font(size=9, bold="利益" in label or "率" in label)
        lc.alignment = _align(h="left")
        lc.border = _border()

        for col, (sc_name, sc), val in zip(range(3, 6), scenarios.items(), vals):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = _fill(sc["fill"])
            c.font = _font(size=10, color=sc["fc"], bold=True)
            c.alignment = _align()
            c.number_format = "0.0%" if is_margin else "#,##0"
            c.border = _border()
        r += 1


# ─────────────────────────────────────────
# Sheet 6 — 資金繰り予測
# ─────────────────────────────────────────
def build_cashflow_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("⑥資金繰り予測")
    ws.sheet_view.showGridLines = False

    col_widths = [2, 20] + [10] * 12 + [12, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    months = ["4月", "5月", "6月", "7月", "8月", "9月",
              "10月", "11月", "12月", "1月", "2月", "3月"]

    _section_title(ws, 2, 2, "  12ヶ月 資金繰り予測（参考）", 14)

    r = 3
    for col, txt in enumerate(["科目"] + months + ["年間合計"], 2):
        _header(ws, r, col, txt)
    r += 1

    pl = data.get("pl", {})
    monthly_rev = (pl.get("revenue") or 0) / 12

    # 簡易的な季節変動係数
    seasonal = [0.85, 0.80, 1.00, 0.95, 1.05, 1.10,
                0.90, 0.85, 1.15, 1.10, 0.95, 1.30]
    total = sum(seasonal)
    seasonal = [s / total * 12 for s in seasonal]

    rows_data = [
        ("売上高（収入）",  [int(monthly_rev * s) for s in seasonal], True,  C["light_blue"]),
        ("仕入・外注費",    [int(monthly_rev * s * 0.60) for s in seasonal], False, C["white"]),
        ("人件費",          [int(monthly_rev * 0.20) for _ in months], False, C["white"]),
        ("販管費（その他）",[int(monthly_rev * 0.08) for _ in months], False, C["white"]),
        ("営業収支差引",    None, True, C["gray_bg"]),
        ("借入返済",        [int((pl.get("long_term_debt") or 0) // 60) for _ in months], False, C["white"]),
        ("月次資金過不足",  None, True, C["amber_bg"]),
        ("期首残高",        None, False, C["white"]),
        ("期末残高",        None, True, C["green_bg"]),
    ]

    cash = (data.get("bs", {}).get("cash") or 0) // 1000

    for i, (label, vals, bold, bg) in enumerate(rows_data):
        lc = ws.cell(row=r, column=2, value=label)
        lc.fill = _fill(bg)
        lc.font = _font(bold=bold, size=9)
        lc.alignment = _align(h="left")
        lc.border = _border()

        if vals:
            total_val = sum(vals)
            for col, v in enumerate(vals, 3):
                c = ws.cell(row=r, column=col, value=v)
                c.fill = _fill(bg)
                c.font = _font(size=9, bold=bold)
                c.alignment = _align()
                c.number_format = "#,##0"
                c.border = _border()
                if label == "売上高（収入）":
                    c.font = _font(size=9, bold=True, color=C["navy"])

            tc = ws.cell(row=r, column=15, value=total_val)
            tc.fill = _fill(C["gray_mid"])
            tc.font = _font(bold=True, size=9)
            tc.alignment = _align()
            tc.number_format = "#,##0"
            tc.border = _border()
        else:
            for col in range(3, 16):
                c = ws.cell(row=r, column=col)
                c.fill = _fill(bg)
                c.font = _font(size=9, italic=True, color=C["gray_dark"])
                c.value = "（自動計算）"
                c.alignment = _align()
                c.border = _border()
        r += 1

    note = ws.cell(row=r + 1, column=2,
                    value="※本シートは概算推計です。実際の資金繰り管理には月次実績との照合が必要です。")
    note.font = _font(size=8, color=C["gray_dark"], italic=True)
    ws.merge_cells(f"B{r+1}:O{r+1}")


# ─────────────────────────────────────────
# メイン生成関数
# ─────────────────────────────────────────
def generate_excel(data: dict, ratios: dict, benchmark: dict,
                   trend_data: list, output_path: str):
    wb = openpyxl.Workbook()

    build_financials_sheet(wb, data)
    build_kpi_sheet(wb, ratios, benchmark)
    build_trend_sheet(wb, trend_data)
    build_benchmark_sheet(wb, ratios, benchmark)
    build_scenario_sheet(wb, data)
    build_cashflow_sheet(wb, data)

    wb.save(output_path)
    print(f"✓ Excel生成完了: {output_path}")
